import pandas as pd
import numpy as np
import openpyxl as xl
import os
import sys
import csv
import xlrd
import shutil

from pathlib import Path
from glob import glob

class ExcelBulk:
    def __init__(self,current_dir,read_folder_name,out_folder):
        self.current_dir = current_dir
        self.read_folder_name = read_folder_name
        self.out_folder = out_folder
        self.out_folder_path = os.path.join(current_dir,out_folder)
        self.read_path = os.path.join(current_dir,read_folder_name)
        self.raw_files = []

    def get_excel_list(self):
        """
        Find all the .xlsx files in the directory and append them to a raw_files list.

        :params:
        self.read_path = the path to the folder where the to-be processed data is.
        :returns:
        raw_files : list of all files ending with '.xlsx' in the read folder.
        """
        raw_files = []
        for file in os.listdir(self.read_path):
            if file.endswith(".xlsx"):
                raw_files.append(file)

        
        print('\nThese are all the .xlsx files inside the folder: \n', raw_files)
        return raw_files

    def get_sheet_names(self,raw_files, n):
        """
        
        Inside the directory list down the number of sheets
        in the nth file name in the raw_files list.

        :params:
        raw_files        : list of all files ending with '.xlsx' in the read folder.

        :returns:
        sheet_names      : All the sheet names in the nth .xlsx file.

        """
        # print(f"Getting Sheet names from the .xlsx files in: {self.read_path}")
        wb1 = xl.load_workbook(os.path.join(self.read_path, raw_files[n]))
        self.sheet_names = wb1.sheetnames
        self.sheet_names.sort()
        print(f"Sheet Names for file {n+1}: {self.sheet_names}")
        wb1.close()
        return self.sheet_names

    def change_sheet_name(self,raw_files,old_name,new_name):
        """
        Changes the name of the sheet in all xlsx files in a folder with the designated name

        :params:
        raw_files : list of all files ending with '.xlsx' in the read folder.
        old_name : List of single variable, which includes the sheet name to be replaced.
        new_name : List of single variable, which includes the new sheet name to be placed.

        :returns:

        """
        for i in range(0, len(raw_files)):
            print(f'Loading Workbook: {raw_files[i]}...')
            loaded_workbook = xl.load_workbook(os.path.join(self.read_path, raw_files[i]))
            print('Loaded Workbook...')
            # printing the sheet names
            print(f'Loading Sheet: {old_name}')

            loaded_sheet = loaded_workbook[old_name]
            loaded_sheet.title = new_name
            print(f"Title {old_name} changed to {new_name}")
            loaded_workbook.save(os.path.join(self.read_path, raw_files[i]))
            loaded_workbook.close()
        return print(i, "files modified")

    def sort_sheets_in_folder_alphabetically(self,raw_files):
        """
        Goes through every file in the raw_files list one by one
        and sorts the existing sheets alphabetically.

        :params:
        raw_files   :   list of all files ending with '.xlsx' in the read folder.

        """
        print(f"\nSorting files in {self.read_path} Alphabetically \n")
        for i in range(0, len(raw_files)):
            wb = xl.load_workbook(os.path.join(self.read_path, raw_files[i]))
            wb._sheets.sort(key=lambda ws: ws.title) # title is the name of the sheet.
            wb.save(os.path.join(self.read_path, raw_files[i]))
            wb.close()
        return

    def remove_sheet(self,raw_files,sheet_to_remove):
        """
        Goes through every .xlsx file one by one and removes the sheet requested.

        :params:
        raw_files       :   list of all files ending with '.xlsx' in the read folder.
        sheet_to_remove : Sheet name to be removed.

        :return:
        The modified .xlsx files are stored as new files in the files_without_sheet
        folder inside the out_folder path.
        """
        print(f"\n Removing {sheet_to_remove} from all .xlsx files.")
        if sheet_to_remove in self.sheet_names:
            self.sheet_names.remove(sheet_to_remove)
        print("Modified Sheet Names list", self.sheet_names)
        for sn in range(0, len(raw_files)):
            wb = xl.load_workbook(f"{self.read_path}/{raw_files[sn]}")
            keep_sheets = self.sheet_names
            for sheetName in wb.sheetnames:
                if sheetName not in keep_sheets:
                    del wb[sheetName]
            files_without_sheet = f"1_without_{sheet_to_remove}"
            path = f"{self.out_folder}//{files_without_sheet}"
            Path(path).mkdir(parents=True, exist_ok=True)
            wb.save(f"{path}//{raw_files[sn]}")
        return files_without_sheet, path


    def multi_file_sheet_to_excel(self,folder_name, raw_files, sheet_number, outputFile):
        """
        Converts all the existing sheets to separate xlsx files

        :params:
        folder_name : the directory to read from
        raw_files   : The list of files in READ_FOLDER_NAME.
        sheet_number: count from 0 to length of raw_files list.
        outputFile  : Name of the output folder where our sheets will be stored.

        :returns:
        new set of conversted .xlsx files with the respected sheet_names.
        """
        # read them in
        os.chdir(folder_name)
        excels = [pd.ExcelFile(name) for name in raw_files]
        # turn them into dataframes
        frames = [x.parse(x.sheet_names[sheet_number], header=None, index_col=None) for x in excels]
        # delete the first row for all frames except the first
        # i.e. remove the header row -- assumes it's the first
        frames[0:] = [df[0:] for df in frames[0:]]
        # concatenate them..
        combined = pd.concat(frames)

        # write it out
        os.chdir("../")
        self.step_2_path = "2_sheet_xlsx"
        Path(self.step_2_path).mkdir(parents=True, exist_ok=True)
        combined.to_excel(f"{self.step_2_path}/{outputFile}", header=False, index=False)
        print(f"{self.sheet_names[sheet_number]} conversion for all files done to {outputFile}")
        return self.step_2_path

    def csv_from_excel(self):
        """
        Converts all xlsx files into their .csv versions.
        Files are stored in Stored in sheet_csv_files.
        """
        self.sheet_csv_files = "3_sheet_csv_files"
        if not os.path.exists(self.sheet_csv_files):
            Path(self.sheet_csv_files).mkdir(parents=True, exist_ok=True)

        for sn in range(0, len(self.sheet_names)):
            wb = xlrd.open_workbook(f'{self.step_2_path}/{self.sheet_names[sn]}.xlsx')
            sh = wb.sheet_by_name('Sheet1')

            csv_file = open(f'{self.sheet_csv_files}//{self.sheet_names[sn]}.csv', 'w', newline='')
            wr = csv.writer(csv_file, delimiter=',', quoting=csv.QUOTE_NONE)
            for rownum in range(sh.nrows):
                wr.writerow(sh.row_values(rownum))
            csv_file.close()
        return self.sheet_csv_files

    def copy_data_and_split(self, num_rows = 20):
        """
        Reads data from the csv files in the sheet_csv_files folder and outputs
        them into the folders in the folder_to_write. This function breaks down
        the CSV files into multiple ones with each file having num_rows per file.

        :params:
        num_rows: Number of rows we want to be copied per file (default = 20)

        """
        # Check whether the specified path is an existing file
        folder_to_write = "4_csv_folders"
        if not os.path.exists(folder_to_write):
            Path(folder_to_write).mkdir(parents=True, exist_ok=True)

        print("\nCreating Folders with names: \n", self.sheet_names)
        for i in range(len(self.sheet_names)):
            if not os.path.exists(f"{folder_to_write}//{self.sheet_names[i]}"):
                os.mkdir(f"{folder_to_write}//{self.sheet_names[i]}")

        print(f"\nConverting single {self.sheet_names} to multiple: \n")
        for gesture in range(len(self.sheet_names)):
            inp_file = "{}//{}.csv".format(self.sheet_csv_files,self.sheet_names[gesture])
            f = "{:{fill}3}"
            folder_name = "{}//{}".format(folder_to_write,self.sheet_names[gesture])
            out_file = "{}_{}.csv".format(self.sheet_names[gesture], f)

            out_file_pattern = os.path.join(folder_name, out_file)
            out_file_pattern = os.path.join(self.out_folder_path, out_file_pattern)
            # print(out_file_pattern)
            max_rows = num_rows

            with open(os.path.join(self.out_folder_path, inp_file), "r") as inp_f:
                reader = csv.reader(inp_f)

                all_rows = []
                cur_file = 1

                for row in reader:
                    all_rows.append(row)
                    if len(all_rows) == max_rows:
                        with open(out_file_pattern.format(cur_file, fill="0"), "w", newline='') as out_f:
                            writer = csv.writer(out_f)
                            writer.writerows(all_rows)
                        all_rows = []
                        cur_file += 1
        return
